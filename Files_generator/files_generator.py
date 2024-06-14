import re
import flet as ft
import time
import os
from openpyxl import load_workbook

global excel_dict
def main(page: ft.Page):
    # t = ft.Text(value="Hello, world!", color="green")
    # page.controls.append(t)
    # page.update()
    # t = ft.Text()
    # page.add(t)  # it's a shortcut for page.controls.append(t) and then page.update()
    #
    # for i in range(10):
    #     t.value = f"Step {i}"
    #     page.update()
    #     time.sleep(1)
    # page.add(
    #     ft.Row(controls=[
    #         ft.TextField(label="Your name"),
    #         ft.ElevatedButton(text="Say my name!")
    #     ])
    # )
    # for i in range(10):
    #     page.controls.append(ft.Text(f"Line {i}"))
    #     if i > 4:
    #         page.controls.pop(0)
    #     page.update()
    #     time.sleep(0.3)
    # def button_clicked(e):
    #     page.add(ft.Text("Clicked!"))
    #
    # page.add(ft.ElevatedButton(text="Click me", on_click=button_clicked))
    # def add_clicked(e):
    #     page.add(ft.Checkbox(label=new_task.value))
    #     new_task.value = ""
    #     new_task.focus()
    #     new_task.update()
    #
    # new_task = ft.TextField(hint_text="Whats needs to be done?", width=300)
    # page.add(ft.Row([new_task, ft.ElevatedButton("Add", on_click=add_clicked)]))
    # first_name = ft.TextField()
    # last_name = ft.TextField()
    # first_name.disabled = True
    # last_name.visible = True
    # page.add(first_name, last_name)
    # first_name = ft.TextField()
    # last_name = ft.TextField()
    # c = ft.Column(controls=[
    #     first_name,
    #     last_name
    # ])
    # c.disabled = True
    # page.add(c)
    # first_name = ft.TextField(label="First name", autofocus=True)
    # last_name = ft.TextField(label="Last name")
    # greetings = ft.Column()
    #
    # def btn_click(e):
    #     greetings.controls.append(ft.Text(f"Hello, {first_name.value} {last_name.value}!"))
    #     first_name.value = ""
    #     last_name.value = ""
    #     page.update()
    #     first_name.focus()
    #
    # page.add(
    #     first_name,
    #     last_name,
    #     ft.ElevatedButton("Say hello!", on_click=btn_click),
    #     greetings,
    # )
    # first_name = ft.Ref[ft.TextField]()
    # last_name = ft.Ref[ft.TextField]()
    # greetings = ft.Ref[ft.Column]()
    #
    # def btn_click(e):
    #     greetings.current.controls.append(
    #         ft.Text(f"Hello, {first_name.current.value} {last_name.current.value}!")
    #     )
    #     first_name.current.value = ""
    #     last_name.current.value = ""
    #     page.update()
    #     first_name.current.focus()
    #
    # page.add(
    #     ft.TextField(ref=first_name, label="First name", autofocus=True),
    #     ft.TextField(ref=last_name, label="Last name"),
    #     ft.ElevatedButton("Say hello!", on_click=btn_click),
    #     ft.Column(ref=greetings),
    # )
    """########################## app ####################################"""
    def generate_files(e):
        folder_path = path_field.value
        files_name = file_name_pattern.value
        file_name_replacement = re.findall("~\[(\w+)]", files_name)
        files_content = file_content_pattern.value
        file_content_replacement = re.findall("~\[(\w+)]", files_content)
        # print(file_content_replacement)
        heads = list(globals()["excel_dict"].keys())
        for value in list(globals()["excel_dict"][heads[0]]):
            with open(os.path.join(folder_path, files_name.replace(f"~[{file_name_replacement[0]}]", value)), "w") as f:
                f.write(files_content)
                # print(value)

    def import_from_excel(e):
        wb = load_workbook(excel_path.value)
        sheet = wb["Sheet1"]
        column_names = [cell.value for cell in sheet[1]]
        data_dict = {}
        for _ in column_names:
            data_dict[_] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for _ in range(len(row)):
                data_dict[list(data_dict.keys())[_]].append(row[_])
        globals()["excel_dict"] = data_dict

    path_field = ft.TextField(label="Folder path", autofocus=True, width=page.width-130)
    file_name_pattern = ft.TextField(label="File name pattern", autofocus=True, width=page.width - 130)
    file_content_pattern = ft.TextField(label="File content pattern", multiline=True, min_lines=6, max_lines=6, width=page.width - 130)
    excel_path = ft.TextField(label="Excel path", autofocus=True, width=page.width-130)

    c = ft.Column(controls=[
        ft.Row(controls=[
            path_field
        ], alignment=ft.MainAxisAlignment.CENTER),
        ft.Row(controls=[
            file_name_pattern
        ], alignment=ft.MainAxisAlignment.CENTER),
        ft.Row(controls=[
            file_content_pattern
        ], alignment=ft.MainAxisAlignment.CENTER),
        ft.Row(controls=[
            excel_path
        ], alignment=ft.MainAxisAlignment.CENTER),
        ft.Row(controls=[
            ft.ElevatedButton("Import data from excel", on_click=import_from_excel),
            ft.ElevatedButton("Generate", on_click=generate_files)
        ],
            alignment=ft.MainAxisAlignment.CENTER),
        ft.Row(controls=[
            ft.Text("""
            How to use:
            1- put the folder where the files will be saved.
            2- put the file name pattern
            3- put the file content pattern
            4- put the path to the excel file where data will be replaced in the patterns is saved
            5- click "Import data from excel"
            6- click "Generate"
            
            PS: the empty parts in the pattern will be filled with data from excel should be in this format:
                   ~[name of the column in the excel file]
            
            """)
        ],
            alignment=ft.MainAxisAlignment.CENTER),
    ])
    page.padding = 20
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.add(c)
    page.update()
    """########################## app ####################################"""


ft.app(target=main)
